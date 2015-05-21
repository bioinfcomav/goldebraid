from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


KEY_CELL_NAME = 'plot_type'
COLUMNS = 'columns'
SCATTER = 'scatter'
XLABEL = 'X-label'
YLABEL = 'Y-label'
TITLE = 'title'
YVALUES = 'Y-values'
XVALUES = 'X-values'
YSTDEV = 'Y-stdev'
XSTDEV = 'X-stdev'

LABELS = (XLABEL, YLABEL, TITLE)
COLUMNS_DATA_TYPE = (YVALUES, XVALUES, YSTDEV)
SCATTER_DATA_TYPE = COLUMNS_DATA_TYPE + (XSTDEV,)


def search_data(ws):
    row_start = None
    row_end = None
    column_start = None
    for row in ws.rows:
        for cell in row:
            cell_content = cell.value
            column_index = cell.column
            row_index = cell.row
            if cell_content == KEY_CELL_NAME:
                row_start = row_index
                column_start = column_index
            if (row_start is not None and column_start == cell.column and
                    cell_content is None):

                row_end = cell.row
                if column_start is None or row_start is None:
                    raise RuntimeError('No data found in excel')
                return (row_start, row_end,
                        column_index_from_string(column_start))
    else:
        row_end = cell.row
        if column_start is None or row_start is None:
            raise RuntimeError('No data found in excel')
        return row_start, row_end, column_index_from_string(column_start)


def parse_xlsx(fpath_or_fhand):

    wb = load_workbook(filename=fpath_or_fhand)
    ws = wb.active
    row_start, row_end, column_start = search_data(ws)
    plot_type = ws.cell(row=row_start, column=column_start + 1).value
    labels = {}
    next_row = row_start + 1
    for row_index in range(next_row, next_row + len(LABELS)):
        key = ws.cell(row=row_index, column=column_start).value
        value = ws.cell(row=row_index, column=column_start + 1).value
        if key not in LABELS:
            raise RuntimeError('Given excel is malformed')
        labels[key] = value
    next_row = row_index + 1

    if plot_type == COLUMNS:
        header_items = COLUMNS_DATA_TYPE
    elif plot_type == SCATTER:
        header_items = SCATTER_DATA_TYPE
    else:
        raise RuntimeError('Can not detect plot type')

    data = {}
    ordered_header = []
    for row_index in range(next_row, row_end + 1):
        for column_index in range(column_start,
                                  column_start + len(header_items)):
            cell_value = ws.cell(row=row_index, column=column_index).value
            if row_index == next_row:
                if cell_value not in header_items:
                    msg = '{} not a valid data type name'
                    raise RuntimeError(msg.format(cell_value))
                data[cell_value] = []
                ordered_header.append(cell_value)
            else:
                data_type = ordered_header[column_index - column_start]
                if data_type in (XSTDEV, YSTDEV, YVALUES):
                    value = float(cell_value)
                elif plot_type == SCATTER and data_type == XVALUES:
                    value = float(cell_value)
                elif plot_type == COLUMNS and data_type == XVALUES:
                    value = str(cell_value)

                data[data_type].append(value)

    return plot_type, labels, data


def draw_columns(labels, data, out_fhand):
    x_vals = data


def plot_from_excel(fpath):
    plot_type, labels, data = parse_xlsx(fpath)
    print data
#     if plot_type == 'columns':
#         draw_columns(labels, data)
#     elif plot_type == 'scatter':
#         draw_scatter(labels, data)
#     else:
#         raise RuntimeError()
def main():
    plot_from_excel('/home/peio/devel/goldenbraid/goldenbraid/tests/data/columns.xlsx')

    plot_from_excel('/home/peio/devel/goldenbraid/goldenbraid/tests/data/scatter.xlsx')
    plot_from_excel('/home/peio/tomato_gbs.xlsx')
if __name__ == '__main__':
    main()
