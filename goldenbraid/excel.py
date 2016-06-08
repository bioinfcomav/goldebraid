# Copyright 2013 Diego Orzaez, Univ.Politecnica Valencia, Consejo Superior de
# Investigaciones Cientificas
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from __future__ import division
import random

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from matplotlib import pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from collections import OrderedDict

MAX_EXPERIMENTS = 3

plt.style.use('ggplot')

KEY_CELL_NAME = 'plot_type'
COLUMNS = 'columns'
SCATTER = 'scatter'
XLABEL = 'X-label'
YLABEL = 'Y-label'
TITLE = 'title'
YVALUES = 'Y-values'
XVALUES = 'X-values'
YSTDEV = 'Y-stdev'
XSTDEV = 'X-stdev'

LABELS = (XLABEL, YLABEL, TITLE)
COLUMNS_DATA_TYPE = (YVALUES, XVALUES, YSTDEV)
SCATTER_DATA_TYPE = COLUMNS_DATA_TYPE + (XSTDEV,)


def search_data(ws):
    row_start = None
    row_end = None
    column_start = None
    for row in ws.rows:
        for cell in row:
            cell_content = cell.value
            column_index = cell.column
            row_index = cell.row
            if cell_content == KEY_CELL_NAME:
                row_start = row_index
                column_start = column_index
            if (row_start is not None and column_start == cell.column and
                    cell_content is None):

                row_end = cell.row - 1
                if column_start is None or row_start is None:
                    raise RuntimeError('No data found in excel')
                return (row_start, row_end,
                        column_index_from_string(column_start))
    else:
        row_end = cell.row
        if column_start is None or row_start is None:
            raise RuntimeError('No data found in excel')
        return row_start, row_end, column_index_from_string(column_start)


def parse_xlsx(fpath_or_fhand):

    wb = load_workbook(filename=fpath_or_fhand)
    ws = wb.active
    row_start, row_end, column_start = search_data(ws)
    plot_type = ws.cell(row=row_start, column=column_start + 1).value
    labels = {}
    next_row = row_start + 1
    for row_index in range(next_row, next_row + len(LABELS)):
        key = ws.cell(row=row_index, column=column_start).value
        value = ws.cell(row=row_index, column=column_start + 1).value
        if key not in LABELS:
            raise RuntimeError('Given excel is malformed')
        labels[key] = value
    next_row = row_index + 1

    if plot_type == COLUMNS:
        header_items = COLUMNS_DATA_TYPE
    elif plot_type == SCATTER:
        header_items = SCATTER_DATA_TYPE
    else:
        raise RuntimeError('Can not detect plot type')

    data = {}
    ordered_header = []
    for row_index in range(next_row, row_end + 1):
        for column_index in range(column_start,
                                  column_start + len(header_items)):
            cell_value = ws.cell(row=row_index, column=column_index).value
            if row_index == next_row:
                if cell_value not in header_items:
                    msg = '{} not a valid data type name'
                    raise RuntimeError(msg.format(cell_value))
                data[cell_value] = []
                ordered_header.append(cell_value)
            else:
                data_type = ordered_header[column_index - column_start]
                if data_type in (XSTDEV, YSTDEV, YVALUES):
                    value = float(cell_value)
                elif plot_type == SCATTER and data_type == XVALUES:
                    value = float(cell_value)
                elif plot_type == COLUMNS and data_type == XVALUES:
                    value = str(cell_value)

                data[data_type].append(value)
    # print plot_type, labels, data
    return plot_type, labels, data


def draw_columns(labels, data, out_fhand):
    canvas, axes, fig = get_canvas_and_axes()
    x_labels = data[XVALUES]
    y_vals = data[YVALUES]
    y_stdev = data.get(YSTDEV, None)
    bar_width = 0.8 if len(x_labels) > 1 else 0.4
    xlabel_pos = [xval + 0.5 for xval in range(len(x_labels))]
    left = [i - (bar_width / 2) for i in xlabel_pos]
    height = y_vals
    kwargs = {}
    if y_stdev:
        kwargs['yerr'] = y_stdev
        kwargs['ecolor'] = 'red'
        capsize = 36 - (len(x_labels) * 4)
        kwargs['capsize'] = 8 if capsize < 8  else capsize

    axes.bar(left=left, height=height, width=bar_width, **kwargs)
    axes.set_xlim(left=0, right=len(x_labels))
    axes.set_title(labels[TITLE])
    axes.set_ylabel(labels[YLABEL])
    axes.set_xlabel(labels[XLABEL])
    axes.set_xticks(xlabel_pos)
    axes.set_xticklabels(data[XVALUES])
    fig.tight_layout()
    canvas.print_figure(out_fhand, format='svg')


def draw_scatter(labels, data, out_fhand):
    canvas, axes, fig = get_canvas_and_axes()
    x_vals = data[XVALUES]
    y_vals = data[YVALUES]
    y_stdev = data.get(YSTDEV, None)
    x_stdev = data.get(XSTDEV, None)
    # axes.scatter(x_vals, y_vals)
    axes.errorbar(x_vals, y_vals, xerr=x_stdev, yerr=y_stdev, fmt='o')

    axes.set_title(labels[TITLE])
    axes.set_ylabel(labels[YLABEL])
    axes.set_xlabel(labels[XLABEL], clip_on=False)
    canvas.print_figure(out_fhand, format='svg')


def plot_from_excel(fpath_or_fhand, out_fhand):
    plot_type, labels, data = parse_xlsx(fpath_or_fhand)
    if plot_type == 'columns':
        draw_columns(labels, data, out_fhand)
    elif plot_type == 'scatter':
        draw_scatter(labels, data, out_fhand)
    else:
        raise RuntimeError()


def get_canvas_and_axes():
    'It returns a matplotlib canvas and axes instance'
    fig = Figure()
    fig.clf()
    canvas = FigureCanvas(fig)
    axes = fig.add_subplot(111)
    return canvas, axes, fig


def _filter_data(data, max_experiments):
    if len(data) < max_experiments:
        return data
    # get max and min
    # only max_experiments
    max_exp = None
    min_exp = None
    prev_max_value = None
    prev_min_value = None
    for exp_name, exp_data in data.items():
        values = exp_data[1]['Y-values']
        max_value = max(values)
        min_value = min(values)
        if prev_max_value is None or max_value > prev_max_value:
            prev_max_value = max_value
            max_exp = exp_name
        if prev_min_value is None or min_value < prev_min_value:
            prev_min_value = min_value
            min_exp = exp_name

    exp_names = data.keys()[:]

    non_filtered_exp = set([min_exp])
    exp_names.pop(exp_names.index(min_exp))
    if max_exp != min_exp:
        non_filtered_exp.add(max_exp)
        exp_names.pop(exp_names.index(max_exp))

    while len(non_filtered_exp) < max_experiments:
        non_filtered_exp.add(random.choice(exp_names))
    return {exp: data[exp] for exp in non_filtered_exp}


def _to_int(val):
    try:
        val = int(val)
    except ValueError:
        pass
    return val


def _prepare_data(data):
    xvalues = set()
    new_data = OrderedDict()
    titles = []
    for exp_name, exp_data in data.items():
        xvalues.update(exp_data[1]['X-values'])
        new_data[exp_name] = {'values': [], 'stdev': []}
        titles.append(exp_data[0]['title'])
    xvalues = sorted(list(xvalues), key=_to_int)
    for exp_name, exp_data in data.items():
        for xvalue in xvalues:
            try:
                index = exp_data[1]['X-values'].index(xvalue)
                yval = exp_data[1]['Y-values'][index]
                try:
                    stdev = exp_data[1]['Y-stdev'][index]
                except IndexError:
                    stdev = 0
            except ValueError:
                yval = 0
                stdev = 0
            new_data[exp_name]['values'].append(yval)
            new_data[exp_name]['stdev'].append(stdev)

    return xvalues, new_data, titles

COMBINED_Y_LABELS = {'SE_001': 'RTA (relative Fluc/Rluc)',
                     'SE_002': 'RTA (relative Fluc/Rluc)',
                     'SE_003': '% transformants',
                     'SE_004': 'recombinant protein yield.',
                     'SE_005': 'overall mutation efficiency (%)'}


def draw_combined_graph(data, out_fhand, exp_type):

    canvas, axes, fig = get_canvas_and_axes()
    axes2 = axes.twiny()
    data = _filter_data(data, max_experiments=MAX_EXPERIMENTS)
    times, data, titles = _prepare_data(data)
    color_scale = ['#80B0BA', '#77F0FD', '#37ACFF', '#426EEA', '#323F9E',
                   '#0F0ADC', '#C974F8', '#8811FF']
    bar_width = 1
    exp_width = bar_width * len(times) + 1
    bar_left_pos = []
    bar_values = []
    bar_stdev = []
    bar_color = []
    xlabel_pos = []
    experiment_labels = data.keys()
    exp_left_pos = []
    exp_color = []
    odd = False
    for exp_index, exp_name in enumerate(experiment_labels):
        exp_values = data[exp_name]['values']
        exp_stdev = data[exp_name]['stdev']
        xlabel_pos.append(exp_index * exp_width + (exp_width / 2))
        exp_left_pos.append(exp_index * exp_width)
        if odd:
            exp_color.append('#F3F3F3')
            odd = False
        else:
            odd = True
            exp_color.append('#E6E6E6')
            # exp_color.append('#464646')

        for bar_index, (val, stdev) in enumerate(zip(exp_values, exp_stdev)):
            bar_stdev.append(stdev)
            bar_values.append(val)
            left_pos = ((exp_index * exp_width) + bar_width * bar_index +
                        bar_width / 2)
            bar_left_pos.append(left_pos)
            bar_color.append(color_scale[bar_index])

    kwargs = {}
    if any(bar_stdev):
        kwargs['yerr'] = bar_stdev
        kwargs['ecolor'] = '#3B383F'
#         capsize = 36 - (len(times) * 4)
#         kwargs['capsize'] = 8 if capsize < 8  else capsize
    rects = axes.bar(left=bar_left_pos, height=bar_values, width=bar_width,
                     color=bar_color, zorder=-1, **kwargs)
    top_lim = axes.get_ylim()[1]
    exp_bars = [top_lim] * len(experiment_labels)
    axes.bar(left=exp_left_pos, height=exp_bars, width=exp_width,
             color=exp_color, zorder=-2)

    axes.set_xticks(xlabel_pos)
    # print(experiment_labels)
    xticklabels = axes.set_xticklabels(experiment_labels)
    for xticklabel in xticklabels:
        url = '/experiment/{}'.format(xticklabel.get_text())
        xticklabel.set_url(url)

    # we have to paint something to get xlabels on top with the same coords
    axes2.bar(left=exp_left_pos, height=exp_bars, width=exp_width,
              alpha=0, color='w')

    # this is only to paint almost invisible bar to put urls on it
    foreground_rects = axes.bar(left=exp_left_pos, height=exp_bars,
                                width=exp_width, alpha=0.1, color='w')
    for frects, experiment_label in zip(foreground_rects, experiment_labels):
        url = '/experiment/{}'.format(experiment_label)
        frects.set_url(url)

    axes.set_xlim(left=0)
    left, right = axes.get_xlim()
    axes2.set_xlim(left=left, right=right)
    axes2.set_xticks(xlabel_pos)
    #axes2.set_xticklabels(titles)
    xticklabels_kwarg = {}
    title_lengths = [len(ti) for ti in titles]
    if title_lengths:
        max_length = max(title_lengths)
    else:
        max_length = 0
    if max_length > 20:
        xticklabels_kwarg = {'rotation': 10}
    axes2.set_xticklabels(titles, **xticklabels_kwarg)

    axes.set_ylabel(COMBINED_Y_LABELS[exp_type])
    axes.legend(rects[:len(times)], times)
    axes2.grid(b=False)
    axes.grid(b=False)
    fig.tight_layout()
    canvas.print_figure(out_fhand, format='svg')

